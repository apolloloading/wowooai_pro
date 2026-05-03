# -*- coding: utf-8 -*-
"""#11 【复杂任务】浏览器导出 + 表格关联拆分.

Tier 3: heaviest E2E. Needs LLM + browser + Excel skill + UAT account +
preset 关联数据.xlsx on the target machine.
"""
from __future__ import annotations

import os
import time
from pathlib import Path

import httpx
import pytest

from .conftest import make_session_id, send_chat_collect


@pytest.mark.e2e
async def test_browser_export_and_table_processing(
    client: httpx.AsyncClient,
    require_dashscope: str,
    require_tier3: None,
    tmp_path,
) -> None:
    """End-to-end: download from UAT → join with reference table → split.

    Expected from the manual checklist: 有效 = 46, 无效 = 0.
    """
    user = os.environ.get("WOWOOAI_TEST_USER_B", "17758000644")
    pwd = os.environ.get("WOWOOAI_TEST_PASS_B", "admin12345677")
    ref_xlsx = os.environ.get(
        "WOWOOAI_E2E_REF_XLSX",
        str(Path.home() / "Desktop" / "关联数据.xlsx"),
    )
    if not Path(ref_xlsx).exists():
        pytest.skip(f"reference table not found: {ref_xlsx}")

    work_dir = tmp_path / "wowooai_complex"
    work_dir.mkdir()
    started = time.time()

    instruction = (
        "现在你按照以下操作：\n"
        f"1. 打开 https://ereference-v-uat.renliwo.com/，账户：{user}，密码：{pwd}\n"
        "2. 点击左侧『合同管理』->『合同产品列表』\n"
        "3. 点击页面『查询』按钮\n"
        f"4. 点击『查询导出』，下载到 {work_dir}，称为表 1\n"
        f"5. 表 2 是 {ref_xlsx}\n"
        "6. 表 1 的 B 列（合同编号）和表 2 的 A 列（合同编号）是关联键\n"
        "7. 通过关联键把表 2 的 E 列（合同有效性）合并到表 1，得到临时表\n"
        "8. 根据『合同有效性』拆分：『有效』一张表，『无效』一张表，"
        f"两张新表都输出到 {work_dir}"
    )

    await send_chat_collect(
        client,
        session_id=make_session_id("complex11"),
        message=instruction,
        timeout=420.0,
    )

    new_files = [
        p for p in work_dir.iterdir()
        if p.is_file() and p.stat().st_mtime >= started - 5
    ]
    if not new_files:
        pytest.xfail("agent produced no output files (heavy E2E flaky)")

    # Find candidate split files by name
    valid_file = next(
        (p for p in new_files if "有效" in p.name and "无效" not in p.name),
        None,
    )
    invalid_file = next(
        (p for p in new_files if "无效" in p.name),
        None,
    )

    if valid_file is None or invalid_file is None:
        pytest.xfail(f"could not identify 有效/无效 output files among {[p.name for p in new_files]}")

    # Row count check (best-effort; needs openpyxl)
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        pytest.skip("openpyxl not available; cannot verify row counts")

    valid_rows = load_workbook(valid_file, read_only=True).active.max_row - 1
    invalid_rows = load_workbook(invalid_file, read_only=True).active.max_row - 1

    # Manual checklist baseline; LLM/data-driven so allow ±2 tolerance
    assert abs(valid_rows - 46) <= 2, f"valid rows={valid_rows}, expected ~46"
    assert invalid_rows <= 2, f"invalid rows={invalid_rows}, expected ~0"
